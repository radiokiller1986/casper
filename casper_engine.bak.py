#!/usr/bin/env python3
"""
CASPER - Causality, Severity & Preventability Evaluation for Research
One-click ADR batch scorer.

Reads an Excel/CSV that uses EITHER the CASPER template columns OR the thesis
ADR-collection variable names (Patient initials, Age, gender, weight_kg, Drug,
adverse events_reaction, adr start date, Rx starting date, Rx stopped, action on
dose, seriousness, relevant Ix, history, outcome, death/hospitalisation,
life threatening/disability ...). Native thesis columns are auto-mapped and a few
fields (temporal, de-challenge, severity flags) are DERIVED from them; anything
that cannot be derived is flagged, never guessed.

Outputs a results workbook: Per-case results, Summary (+charts), Data flags.

Usage:
    python casper_engine.py  <input.xlsx|csv>  [output.xlsx]

Author: Puneet Paliwal. Licensed under Apache-2.0.
"""
import sys, os, re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

NAVY="1F3864"; BLUE="2E5496"; GREEN="C6E0B4"; AMBER="FFE699"; RED="F4B6B6"; GREY="F2F2F2"
F="Arial"

# ---------- column aliasing: native header (lowercased, squashed) -> canonical ----------
ALIAS={
 "patients initials":"patient_id","patient initials":"patient_id","patient initial":"patient_id",
 "age":"age_years","age years":"age_years","age_years":"age_years",
 "gender":"sex","sex":"sex",
 "weight":"weight_kg","weight kg":"weight_kg","weight_kg":"weight_kg","weight in kg":"weight_kg",
 "drug":"suspected_drug","suspected drug":"suspected_drug","suspected_drug":"suspected_drug","medication":"suspected_drug",
 "adverse events reaction":"adr_description","adverse event reaction":"adr_description","adverse event":"adr_description",
 "reaction":"adr_description","adr":"adr_description","adr_description":"adr_description","adverse events_reaction":"adr_description",
 "adr star date":"adr_onset_date","adr start date":"adr_onset_date","reaction start date":"adr_onset_date","adr_onset_date":"adr_onset_date",
 "rx starting date":"therapy_start_date","rx start date":"therapy_start_date","therapy started":"therapy_start_date",
 "date started":"therapy_start_date","therapy_start_date":"therapy_start_date","rx_start":"therapy_start_date",
 "rx stopped":"therapy_stop_date","therapy stopped":"therapy_stop_date","date stopped":"therapy_stop_date","therapy_stop_date":"therapy_stop_date",
 "indication":"indication",
 "dose":"dose","route":"route","frequency":"frequency",
 "action on dose":"action_on_dose","action taken":"action_on_dose","action on dose as per c":"action_on_dose","action_on_dose":"action_on_dose",
 "seriousness y n":"seriousness","seriousness":"seriousness","seriousness yn":"seriousness","serious":"seriousness",
 "relevant ix if any":"investigations","relevant ix":"investigations","relevant investigations":"investigations","investigations":"investigations",
 "history":"comorbidities","relevant medical history":"comorbidities","medical history":"comorbidities","comorbidities":"comorbidities",
 "outcome":"adr_outcome","adr_outcome":"adr_outcome",
 "death hospitalisation":"death_hosp","death/hospitalisation":"death_hosp","death hospitalization":"death_hosp","death_hosp":"death_hosp",
 "life threatning disability":"lifethreat_disab","life threatening disability":"lifethreat_disab",
 "life threatning/disability":"lifethreat_disab","life threatening/disability":"lifethreat_disab","lifethreat_disab":"lifethreat_disab",
 "case id":"case_id","case_id":"case_id","sl no":"sl_no","s no":"sl_no","sr no":"sl_no",
 "liver injury":"liver_injury","dili":"liver_injury","drug induced liver injury":"liver_injury","liver_injury":"liver_injury",
 "lab pattern":"lab_pattern","pattern":"lab_pattern","alt":"alt_value","alt value":"alt_value","alt uln":"alt_uln",
 "alp":"alp_value","alp value":"alp_value","alp uln":"alp_uln","alcohol":"alcohol_use","alcohol use":"alcohol_use","pregnancy":"pregnancy",
}
# canonical scoring columns kept as-is if already present
PASS_THROUGH={"case_id","patient_id","age_years","sex","weight_kg","comorbidities","suspected_drug","dose","route",
 "frequency","indication","therapy_start_date","therapy_stop_date","concomitant_drugs","adr_description","meddra_pt",
 "adr_onset_date","adr_outcome","seriousness","investigations","action_on_dose","death_hosp","lifethreat_disab",
 "c_prev_reports","c_temporal","c_dechallenge","c_rechallenge","c_alt_causes","c_placebo","c_toxic_level",
 "c_dose_response","c_prior_exposure","c_objective_confirm","c_event_definitive",
 "s_treatment_change","s_antidote_required","s_caused_admission","s_increased_los","s_los_days","s_intensive_care",
 "s_permanent_harm","s_contributed_death",
 "p_a1_allergy_history","p_a2_drug_inappropriate","p_a3_dose_inappropriate","p_a4_toxic_level","p_a5_known_treatment",
 "p_b1_monitoring_missed","p_b2_interaction","p_b3_poor_compliance","p_b4_preventive_missed",
 "liver_injury","lab_pattern","alt_value","alt_uln","alp_value","alp_uln","rucam_time_to_onset_days",
 "rucam_course","alcohol_use","pregnancy","rucam_concomitant","rucam_nondrug_causes","rucam_prev_info","rucam_rechallenge"}

def _norm(h):
    h=str(h).strip().lower()
    h=re.sub(r"[._/]+"," ",h)
    h=re.sub(r"[^a-z0-9 ]+","",h)
    h=re.sub(r"\s+"," ",h).strip()
    return h

def normalize_columns(df):
    ren={}
    for c in df.columns:
        n=_norm(c)
        if c in PASS_THROUGH: continue
        if n in ALIAS: ren[c]=ALIAS[n]
        elif n in PASS_THROUGH: ren[c]=n
    df=df.rename(columns=ren)
    # collapse duplicate columns (keep first non-empty)
    df=df.loc[:,~df.columns.duplicated()]
    return df

def _contains(v,*subs):
    s=str(v).lower()
    return any(x in s for x in subs)

def _to_date(v):
    try: return pd.to_datetime(v, dayfirst=True, errors="coerce")
    except Exception: return pd.NaT

def derive_fields(df):
    """Fill canonical scoring columns from native thesis columns where missing."""
    def has(col): return col in df.columns
    for col in ["c_temporal","c_dechallenge","s_treatment_change","s_antidote_required","s_caused_admission",
                "s_increased_los","s_intensive_care","s_permanent_harm","s_contributed_death","case_id"]:
        if not has(col): df[col]=""
    for i,row in df.iterrows():
        # case_id fallback
        if not str(row.get("case_id","")).strip() or str(row.get("case_id")).lower()=="nan":
            sid=row.get("sl_no","") if has("sl_no") else ""
            df.at[i,"case_id"]= (f"CASE-{str(sid).split('.')[0]}" if str(sid).strip() and str(sid).lower()!="nan"
                                 else f"CASE-{i+1}")
        # temporal from dates (only if not explicitly provided)
        if not str(row.get("c_temporal","")).strip():
            sd=_to_date(row.get("therapy_start_date")); od=_to_date(row.get("adr_onset_date"))
            if pd.notna(sd) and pd.notna(od):
                df.at[i,"c_temporal"]= "Reasonable" if od>=sd else "Improbable"
        # dechallenge from action on dose + outcome
        if not str(row.get("c_dechallenge","")).strip() and has("action_on_dose"):
            act=str(row.get("action_on_dose","")).lower(); out=str(row.get("adr_outcome","")).lower()
            if _contains(act,"withdraw","stopp","reduc","discontinu","held"):
                df.at[i,"c_dechallenge"]= "Improved" if _contains(out,"recover","resolv","improv") else "No change"
            elif _contains(act,"not changed","continued","unchanged","increase"):
                df.at[i,"c_dechallenge"]="No change"
        # treatment change (Hartwig L1/L2) from action on dose
        if not str(row.get("s_treatment_change","")).strip() and has("action_on_dose"):
            act=str(row.get("action_on_dose","")).lower()
            if _contains(act,"withdraw","stopp","reduc","discontinu","held","changed"):
                df.at[i,"s_treatment_change"]="Drug held-stopped-changed"
            elif _contains(act,"not changed","continued","unchanged"):
                df.at[i,"s_treatment_change"]="No change"
        # severity flags from outcome / death_hosp / lifethreat columns
        out=str(row.get("adr_outcome","")).lower()
        dh = str(row.get("death_hosp","")).lower() if has("death_hosp") else ""
        lt = str(row.get("lifethreat_disab","")).lower() if has("lifethreat_disab") else ""
        if not str(row.get("s_contributed_death","")).strip():
            if _contains(out,"fatal","death","died") or _contains(dh,"death","died","fatal"):
                df.at[i,"s_contributed_death"]="Yes"
        if not str(row.get("s_permanent_harm","")).strip():
            if _contains(out,"sequelae") or _contains(lt,"disab","permanent"):
                df.at[i,"s_permanent_harm"]="Yes"
        if not str(row.get("s_intensive_care","")).strip():
            if _contains(lt,"life") or _contains(out,"life threat"):
                df.at[i,"s_intensive_care"]="Yes"
        if not str(row.get("s_caused_admission","")).strip():
            if _contains(dh,"hosp","admit","admission"):
                df.at[i,"s_caused_admission"]="Yes"
        if not str(row.get("s_increased_los","")).strip():
            if _contains(dh,"prolong","prolonged los","increase"):
                df.at[i,"s_increased_los"]="Yes"
        # if we had ANY severity source column, treat un-set flags as explicit "No"
        sev_source = any([out.strip(), dh.strip(), lt.strip(),
                          str(row.get("action_on_dose","")).strip()])
        if sev_source:
            for k in ["s_contributed_death","s_permanent_harm","s_intensive_care",
                      "s_caused_admission","s_increased_los","s_antidote_required"]:
                if not str(df.at[i,k]).strip() if k in df.columns else True:
                    if k not in df.columns: df[k]=""
                    if not str(df.at[i,k]).strip(): df.at[i,k]="No"
            if not str(df.at[i,"s_treatment_change"]).strip():
                df.at[i,"s_treatment_change"]="No change"
    return df

def g(row, key, default=""):
    v=row.get(key, default)
    if pd.isna(v): return default
    return str(v).strip()

# ---------------- NARANJO ----------------
def naranjo(row):
    miss=[]; s=0
    s+= 1 if g(row,"c_prev_reports")=="Yes" else 0
    v=g(row,"c_temporal")
    if v in ("Plausible","Reasonable"): s+=2
    elif v in ("Improbable","None"): s+=-1
    else: miss.append("c_temporal")
    if g(row,"c_dechallenge")=="Improved": s+=1
    elif g(row,"c_dechallenge")=="": miss.append("c_dechallenge")
    v=g(row,"c_rechallenge")
    if v=="Reaction recurred": s+=2
    elif v=="No recurrence": s+=-1
    # "Not done"/blank = no information -> 0 points, not a data gap
    v=g(row,"c_alt_causes")
    if v=="Absent": s+=2
    elif v in ("Present","Possible"): s+=-1
    elif v=="": miss.append("c_alt_causes")
    v=g(row,"c_placebo")
    if v=="Reaction present": s+=-1
    elif v=="Reaction absent": s+=1
    s+= 1 if g(row,"c_toxic_level")=="Yes" else 0
    s+= 1 if g(row,"c_dose_response")=="Yes" else 0
    s+= 1 if g(row,"c_prior_exposure")=="Yes" else 0
    s+= 1 if g(row,"c_objective_confirm")=="Yes" else 0
    cat="Definite" if s>=9 else "Probable" if s>=5 else "Possible" if s>=1 else "Doubtful"
    return s, cat, miss

# ---------------- WHO-UMC ----------------
def whoumc(row):
    t=g(row,"c_temporal"); alt=g(row,"c_alt_causes")
    de=g(row,"c_dechallenge"); re_=g(row,"c_rechallenge"); defin=g(row,"c_event_definitive")
    if t=="" or alt=="" or de=="": return "Conditional / Unclassified"
    if t=="Improbable": return "Unlikely"
    if t=="None": return "Unassessable"
    if t=="Plausible" and alt=="Absent" and de=="Improved" and (re_=="Reaction recurred" or defin=="Yes"):
        return "Certain"
    if t in ("Plausible","Reasonable") and alt=="Absent" and de in ("Improved","No change"):
        return "Probable / Likely"
    return "Possible"

# ---------------- HARTWIG & SIEGEL ----------------
def hartwig(row):
    keys=["s_contributed_death","s_permanent_harm","s_intensive_care","s_caused_admission",
          "s_increased_los","s_antidote_required","s_treatment_change"]
    known=[k for k in keys if g(row,k)!=""]
    miss=[] if known else ["severity inputs"]
    def y(k): return g(row,k)=="Yes"
    if y("s_contributed_death"): return 7,"Severe",miss
    if y("s_permanent_harm"): return 6,"Severe",miss
    if y("s_intensive_care"): return 5,"Severe",miss
    if y("s_caused_admission") or y("s_increased_los"): return 4,"Moderate",miss
    if y("s_antidote_required"): return 3,"Moderate",miss
    if g(row,"s_treatment_change")=="Drug held-stopped-changed": return 2,"Mild",miss
    return 1,"Mild",miss

# ---------------- SCHUMOCK & THORNTON ----------------
def preventability(row):
    A=["p_a1_allergy_history","p_a2_drug_inappropriate","p_a3_dose_inappropriate","p_a4_toxic_level","p_a5_known_treatment"]
    B=["p_b1_monitoring_missed","p_b2_interaction","p_b3_poor_compliance","p_b4_preventive_missed"]
    filled=[k for k in A+B if g(row,k)!=""]
    miss=[] if filled else ["preventability inputs"]
    if any(g(row,k)=="Yes" for k in A): return "Definitely preventable", miss
    if any(g(row,k)=="Yes" for k in B): return "Probably preventable", miss
    return "Not preventable", miss

# ---------------- RUCAM (drug-induced liver injury) ----------------
def rucam(row):
    if g(row,"liver_injury")!="Yes":
        return "", "Not a liver case"
    flags=[]
    pattern=g(row,"lab_pattern")
    try:
        alt=float(g(row,"alt_value") or 0); altu=float(g(row,"alt_uln") or 0)
        alp=float(g(row,"alp_value") or 0); alpu=float(g(row,"alp_uln") or 0)
        if altu>0 and alpu>0 and alp>0:
            R=(alt/altu)/(alp/alpu)
            pattern="Hepatocellular" if R>=5 else "Cholestatic" if R<=2 else "Mixed"
    except Exception:
        pass
    if not pattern:
        pattern="Hepatocellular"; flags.append("lab_pattern assumed")
    hep=(pattern=="Hepatocellular")
    sc=0
    # 1. time to onset
    t=g(row,"rucam_time_to_onset_days")
    if not t:
        sd=_to_date(row.get("therapy_start_date")); od=_to_date(row.get("adr_onset_date"))
        if pd.notna(sd) and pd.notna(od): t=(od-sd).days
    try: t=float(t)
    except Exception: t=None
    if t is None: flags.append("time-to-onset")
    else: sc += 2 if 5<=t<=90 else 1
    # 2. course (ALT for hepatocellular, ALP for cholestatic/mixed)
    course=g(row,"rucam_course")
    if hep:
        cmap={"Decrease >=50% in 8d":3,"Decrease >=50% in 30d":2,"Decrease >=50% after 30d":0,
              "Decrease <50% or re-rise":-2,"No info / drug continued":0,"Persists or rises":-2}
    else:
        cmap={"Decrease >=50% in 8d":2,"Decrease >=50% in 30d":2,"Decrease >=50% in 180d":2,
              "Decrease <50% or re-rise":1,"No info / drug continued":0,"Persists or rises":0}
    sc += cmap.get(course,0)
    # 3. risk factors
    try: age=float(g(row,"age_years") or 0)
    except Exception: age=0
    sc += 1 if age>=55 else 0
    if hep:
        sc += 1 if g(row,"alcohol_use")=="Yes" else 0
    else:
        sc += 1 if (g(row,"alcohol_use")=="Yes" or g(row,"pregnancy")=="Yes") else 0
    # 4. concomitant drugs
    con=g(row,"rucam_concomitant")
    if not con:
        con = "None" if not g(row,"concomitant_drugs") else "Compatible timing"
    sc += {"None":0,"Incompatible timing":0,"Compatible timing":-1,
           "Known hepatotoxin compatible":-2,"Proven role (rechallenge+)":-3}.get(con,0)
    # 5. search for non-drug causes
    nd=g(row,"rucam_nondrug_causes")
    if not nd:
        nd={"Absent":"6 of Group I ruled out","Possible":"4-5 of Group I ruled out",
            "Present":"Non-drug cause probable","Unknown":"4-5 of Group I ruled out"}.get(g(row,"c_alt_causes"),"4-5 of Group I ruled out")
    sc += {"All (Group I+II) ruled out":2,"6 of Group I ruled out":1,"4-5 of Group I ruled out":0,
           "<4 of Group I ruled out":-2,"Non-drug cause probable":-3}.get(nd,0)
    # 6. previous hepatotoxicity information
    pi=g(row,"rucam_prev_info")
    if not pi:
        pi="Labelled in product info" if g(row,"c_prev_reports")=="Yes" else "Reaction unknown"
    sc += {"Labelled in product info":2,"Published but unlabelled":1,"Reaction unknown":0}.get(pi,0)
    # 7. response to readministration
    rc=g(row,"rucam_rechallenge")
    if not rc:
        rc={"Reaction recurred":"Positive","No recurrence":"Negative","Not done":"Not done / NA"}.get(g(row,"c_rechallenge"),"Not done / NA")
    sc += {"Positive":3,"Compatible":1,"Negative":-2,"Not done / NA":0}.get(rc,0)
    cat=("Highly probable" if sc>=9 else "Probable" if sc>=6 else "Possible" if sc>=3
         else "Unlikely" if sc>=1 else "Excluded")
    return sc, f"{cat} ({pattern})"

# ---------------- LOAD / SCORE / REPORT ----------------
def load_rows(path):
    if path.lower().endswith(".csv"):
        df=pd.read_csv(path, dtype=str, keep_default_na=False)
    else:
        xls=pd.ExcelFile(path)
        sheet=next((s for s in ["Upload template","Sheet1"] if s in xls.sheet_names), xls.sheet_names[0])
        df=pd.read_excel(path, sheet_name=sheet, dtype=str)
    df=normalize_columns(df)
    # keep genuine rows: must have a suspected drug
    if "suspected_drug" not in df.columns:
        raise SystemExit("No 'Drug'/'suspected_drug' column found after mapping.")
    drug=df["suspected_drug"].astype(str).str.strip()
    df=df[(drug!="") & (drug.str.lower()!="nan")].reset_index(drop=True)
    df=derive_fields(df)
    return df

def score(df):
    out=[]
    for _,row in df.iterrows():
        ns,nc,nm=naranjo(row); who=whoumc(row); hl,hg,hm=hartwig(row); pv,pm=preventability(row)
        rsc,rcat=rucam(row)
        flags=[]
        if nm or who=="Conditional / Unclassified": flags.append("causality")
        if hm: flags.append("severity")
        if pm: flags.append("preventability")
        out.append({
            "case_id":g(row,"case_id"),"patient_id":g(row,"patient_id"),
            "suspected_drug":g(row,"suspected_drug"),"adr_description":g(row,"adr_description"),
            "Naranjo score":ns,"Naranjo causality":nc,"WHO-UMC causality":who,
            "Hartwig level":hl,"Severity grade":hg,"Preventability":pv,
            "RUCAM score":rsc,"RUCAM causality":rcat,
            "Data flags":"; ".join(flags) if flags else "complete",
            "Missing/assumed fields":"; ".join(sorted(set(nm+hm+pm))),
        })
    return pd.DataFrame(out)

def write_report(res, out_path):
    wb=Workbook(); thin=Side(style="thin",color="BFBFBF"); bord=Border(thin,thin,thin,thin)
    hf=Font(name=F,bold=True,color="FFFFFF"); hfill=PatternFill("solid",fgColor=BLUE)
    ws=wb.active; ws.title="Per-case results"; cols=list(res.columns)
    for j,c in enumerate(cols,1):
        cell=ws.cell(1,j,c); cell.font=hf; cell.fill=hfill; cell.border=bord
        cell.alignment=Alignment(wrap_text=True,horizontal="center",vertical="center")
    for i,(_,r) in enumerate(res.iterrows(),2):
        for j,c in enumerate(cols,1):
            cell=ws.cell(i,j,r[c]); cell.border=bord; cell.font=Font(name=F,size=10)
            cell.alignment=Alignment(vertical="center",wrap_text=True)
            if c=="Severity grade":
                cell.fill=PatternFill("solid",fgColor={"Mild":GREEN,"Moderate":AMBER,"Severe":RED}.get(r[c],"FFFFFF"))
            if c=="Data flags" and r[c]!="complete":
                cell.fill=PatternFill("solid",fgColor=AMBER)
    for j,w in enumerate([12,11,16,28,11,15,18,9,12,20,10,22,14,34][:len(cols)],1):
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=40

    sm=wb.create_sheet("Summary")
    sm["A1"]="CASPER — Aggregate summary"; sm["A1"].font=Font(name=F,bold=True,size=14,color=NAVY)
    sm["A2"]=f"Total cases scored: {len(res)}"; sm["A2"].font=Font(name=F,bold=True)
    r=4
    def block(title, series, order=None):
        nonlocal r
        sm.cell(r,1,title).font=Font(name=F,bold=True,color="FFFFFF"); sm.cell(r,1).fill=PatternFill("solid",fgColor=NAVY)
        sm.cell(r,2).fill=PatternFill("solid",fgColor=NAVY); r+=1
        sm.cell(r,1,"Category").font=Font(name=F,bold=True); sm.cell(r,2,"n").font=Font(name=F,bold=True)
        sm.cell(r,3,"%").font=Font(name=F,bold=True); r+=1
        vc=series.value_counts(); cats=order if order else list(vc.index); start=r; tot=len(series)
        for cat in cats:
            n=int(vc.get(cat,0)); sm.cell(r,1,cat); sm.cell(r,2,n)
            sm.cell(r,3, f"{(100*n/tot):.1f}%" if tot else "0%"); r+=1
        ch=BarChart(); ch.type="bar"; ch.title=title; ch.height=5; ch.width=11
        ch.add_data(Reference(sm,min_col=2,min_row=start,max_row=r-1))
        ch.set_categories(Reference(sm,min_col=1,min_row=start,max_row=r-1)); ch.legend=None
        sm.add_chart(ch, f"E{start-1}"); r+=1
    block("Causality — Naranjo", res["Naranjo causality"], ["Definite","Probable","Possible","Doubtful"])
    block("Causality — WHO-UMC", res["WHO-UMC causality"])
    block("Severity — Hartwig & Siegel", res["Severity grade"], ["Mild","Moderate","Severe"])
    block("Preventability — Schumock & Thornton", res["Preventability"], ["Definitely preventable","Probably preventable","Not preventable"])
    sm.column_dimensions["A"].width=30; sm.column_dimensions["B"].width=8; sm.column_dimensions["C"].width=10

    fl=wb.create_sheet("Data flags"); flagged=res[res["Data flags"]!="complete"]
    fl["A1"]="Rows needing reviewer input (inputs missing or assumed)"; fl["A1"].font=Font(name=F,bold=True,size=12,color=NAVY)
    if len(flagged)==0:
        fl["A3"]="None — all rows had complete inputs."; fl["A3"].font=Font(name=F,italic=True)
    else:
        hdr=["case_id","suspected_drug","Data flags","Missing/assumed fields"]
        for j,c in enumerate(hdr,1):
            cell=fl.cell(3,j,c); cell.font=hf; cell.fill=hfill; cell.border=bord
        for i,(_,r2) in enumerate(flagged.iterrows(),4):
            for j,c in enumerate(hdr,1):
                cell=fl.cell(i,j,r2[c]); cell.border=bord; cell.font=Font(name=F,size=10); cell.alignment=Alignment(wrap_text=True,vertical="top")
        for j,w in enumerate([14,18,22,55],1): fl.column_dimensions[get_column_letter(j)].width=w
    wb.save(out_path)

def main():
    if len(sys.argv)<2:
        print("Usage: python casper_engine.py <input.xlsx|csv> [output.xlsx]"); sys.exit(1)
    inp=sys.argv[1]
    out=sys.argv[2] if len(sys.argv)>2 else os.path.splitext(inp)[0].replace("_template","").replace("_upload","")+"_RESULTS.xlsx"
    df=load_rows(inp); res=score(df); write_report(res,out)
    print(f"Scored {len(res)} case(s).  ->  {out}")
    print(res[["case_id","Naranjo causality","WHO-UMC causality","Severity grade","Preventability","Data flags"]].to_string(index=False))

if __name__=="__main__":
    main()
