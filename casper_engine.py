#!/usr/bin/env python3
"""
CASPER - Causality, Severity & Preventability Evaluation for Research
One-click ADR batch scorer.

Reads an Excel/CSV that uses EITHER the CASPER template columns OR the thesis
ADR-collection variable names (Patient initials, Age, gender, weight_kg, Drug,
adverse events_reaction, adr start date, Rx starting date, Rx stopped, action on
dose, seriousness, relevant Ix, history, outcome, death/hospitalisation,
life threatening/disability ...). Native thesis columns are auto-mapped and a few
fields (temporal, de-challenge, severity flags) are DERIVED from them; anything
that cannot be derived is flagged, never guessed.

Outputs a results workbook: Per-case results, Summary (+charts), Data flags.

Usage:
    python casper_engine.py  <input.xlsx|csv>  [output.xlsx]

Author: Puneet Paliwal. Licensed under Apache-2.0.
"""
import sys, os, re, hashlib, datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

NAVY="1F3864"; BLUE="2E5496"; GREEN="C6E0B4"; AMBER="FFE699"; RED="F4B6B6"; GREY="F2F2F2"
CASPER_VERSION="0.2.0"
CASPER_DOI="10.5281/zenodo.20646452"   # replace with your Zenodo DOI after first release
CASPER_URL="https://github.com/radiokiller1986/casper"
CASPER_CITATION=(f"Paliwal P. CASPER: Causality, Severity & Preventability Evaluation for Research "
                 f"(v{CASPER_VERSION}). 2026. DOI: {CASPER_DOI}. {CASPER_URL}")

F="Arial"

# ---------- column aliasing: native header (lowercased, squashed) -> canonical ----------
ALIAS={
 "patients initials":"patient_id","patient initials":"patient_id","patient initial":"patient_id",
 "age":"age_years","age years":"age_years","age_years":"age_years",
 "gender":"sex","sex":"sex",
 "weight":"weight_kg","weight kg":"weight_kg","weight_kg":"weight_kg","weight in kg":"weight_kg",
 "drug":"suspected_drug","suspected drug":"suspected_drug","suspected_drug":"suspected_drug","medication":"suspected_drug",
 "adverse events reaction":"adr_description","adverse event reaction":"adr_description","adverse event":"adr_description",
 "reaction":"adr_description","adr":"adr_description","adr_description":"adr_description","adverse events_reaction":"adr_description",
 "adr star date":"adr_onset_date","adr start date":"adr_onset_date","reaction start date":"adr_onset_date","adr_onset_date":"adr_onset_date",
 "rx starting date":"therapy_start_date","rx start date":"therapy_start_date","therapy started":"therapy_start_date",
 "date started":"therapy_start_date","therapy_start_date":"therapy_start_date","rx_start":"therapy_start_date",
 "rx stopped":"therapy_stop_date","therapy stopped":"therapy_stop_date","date stopped":"therapy_stop_date","therapy_stop_date":"therapy_stop_date",
 "indication":"indication",
 "dose":"dose","route":"route","frequency":"frequency",
 "action on dose":"action_on_dose","action taken":"action_on_dose","action on dose as per c":"action_on_dose","action_on_dose":"action_on_dose",
 "seriousness y n":"seriousness","seriousness":"seriousness","seriousness yn":"seriousness","serious":"seriousness",
 "relevant ix if any":"investigations","relevant ix":"investigations","relevant investigations":"investigations","investigations":"investigations",
 "history":"comorbidities","relevant medical history":"comorbidities","medical history":"comorbidities","comorbidities":"comorbidities",
 "outcome":"adr_outcome","adr_outcome":"adr_outcome",
 "death hospitalisation":"death_hosp","death/hospitalisation":"death_hosp","death hospitalization":"death_hosp","death_hosp":"death_hosp",
 "life threatning disability":"lifethreat_disab","life threatening disability":"lifethreat_disab",
 "life threatning/disability":"lifethreat_disab","life threatening/disability":"lifethreat_disab","lifethreat_disab":"lifethreat_disab",
 "case id":"case_id","case_id":"case_id","sl no":"sl_no","s no":"sl_no","sr no":"sl_no",
}
# canonical scoring columns kept as-is if already present
PASS_THROUGH={"case_id","patient_id","age_years","sex","weight_kg","comorbidities","suspected_drug","dose","route",
 "frequency","indication","therapy_start_date","therapy_stop_date","concomitant_drugs","adr_description","meddra_pt",
 "adr_onset_date","adr_outcome","seriousness","investigations","action_on_dose","death_hosp","lifethreat_disab",
 "c_prev_reports","c_temporal","c_dechallenge","c_rechallenge","c_alt_causes","c_placebo","c_toxic_level",
 "c_dose_response","c_prior_exposure","c_objective_confirm","c_event_definitive",
 "s_treatment_change","s_antidote_required","s_caused_admission","s_increased_los","s_los_days","s_intensive_care",
 "s_permanent_harm","s_contributed_death",
 "p_a1_allergy_history","p_a2_drug_inappropriate","p_a3_dose_inappropriate","p_a4_toxic_level","p_a5_known_treatment",
 "p_b1_monitoring_missed","p_b2_interaction","p_b3_poor_compliance","p_b4_preventive_missed",
 "seriousness","e2a_life_threatening","e2a_congenital_anomaly","e2a_important_medical_event"}

def _norm(h):
    h=str(h).strip().lower()
    h=re.sub(r"[._/]+"," ",h)
    h=re.sub(r"[^a-z0-9 ]+","",h)
    h=re.sub(r"\s+"," ",h).strip()
    return h

def normalize_columns(df):
    ren={}
    for c in df.columns:
        n=_norm(c)
        if c in PASS_THROUGH: continue
        if n in ALIAS: ren[c]=ALIAS[n]
        elif n in PASS_THROUGH: ren[c]=n
    df=df.rename(columns=ren)
    # collapse duplicate columns (keep first non-empty)
    df=df.loc[:,~df.columns.duplicated()]
    return df

def _contains(v,*subs):
    s=str(v).lower()
    return any(x in s for x in subs)

_DAYFIRST=True  # set by detect_date_format()

def _to_date(v):
    try: return pd.to_datetime(v, dayfirst=_DAYFIRST, errors="coerce")
    except Exception: return pd.NaT

def detect_date_format(df):
    """Decide dd/mm vs mm/dd for the whole dataset by scanning for unambiguous tokens."""
    global _DAYFIRST
    import re as _re
    dayfirst_votes=0; monthfirst_votes=0
    for col in ["therapy_start_date","therapy_stop_date","adr_onset_date"]:
        if col not in df.columns: continue
        for v in df[col].dropna().astype(str):
            m=_re.match(r"^(\d{1,2})[\/\-.](\d{1,2})[\/\-.]\d{2,4}$", v.strip())
            if not m: continue
            a,b=int(m.group(1)),int(m.group(2))
            if a>12 and b<=12: dayfirst_votes+=1      # first token must be a day
            elif b>12 and a<=12: monthfirst_votes+=1  # second token must be a day -> mm/dd
    _DAYFIRST = not (monthfirst_votes>dayfirst_votes)
    return _DAYFIRST

def derive_fields(df):
    """Fill canonical scoring columns from native thesis columns where missing."""
    def has(col): return col in df.columns
    for col in ["c_temporal","c_dechallenge","s_treatment_change","s_antidote_required","s_caused_admission",
                "s_increased_los","s_intensive_care","s_permanent_harm","s_contributed_death","case_id"]:
        if not has(col): df[col]=""
    if "data_integrity" not in df.columns: df["data_integrity"]=""
    for i,row in df.iterrows():
        # case_id fallback
        if not str(row.get("case_id","")).strip() or str(row.get("case_id")).lower()=="nan":
            sid=row.get("sl_no","") if has("sl_no") else ""
            df.at[i,"case_id"]= (f"CASE-{str(sid).split('.')[0]}" if str(sid).strip() and str(sid).lower()!="nan"
                                 else f"CASE-{i+1}")
        # timeline coherence
        sd=_to_date(row.get("therapy_start_date")); od=_to_date(row.get("adr_onset_date"))
        coherent = (pd.isna(sd) or pd.isna(od) or od>=sd)
        integ=[]
        if pd.notna(sd) and pd.notna(od) and od<sd:
            integ.append("reaction onset precedes drug start")
        try:
            ag=float(str(row.get("age_years","")).strip() or "nan")
            if ag<0 or ag>120: integ.append("implausible age")
        except Exception: pass
        if integ: df.at[i,"data_integrity"]="; ".join(integ)
        # temporal from dates
        if not str(row.get("c_temporal","")).strip() and pd.notna(sd) and pd.notna(od):
            df.at[i,"c_temporal"]= "Reasonable" if od>=sd else "Improbable"
        # dechallenge from action on dose + outcome -- ONLY when timeline is coherent
        if not str(row.get("c_dechallenge","")).strip() and has("action_on_dose") and coherent:
            act=str(row.get("action_on_dose","")).lower(); out=str(row.get("adr_outcome","")).lower()
            neg=_contains(out,"not recover","not resolv","unresolved","ongoing","fatal","death","not improv")
            improved=(not neg) and _contains(out,"recover","resolv","improv")
            if _contains(act,"withdraw","stopp","reduc","discontinu","held"):
                df.at[i,"c_dechallenge"]= "Improved" if improved else "No change"
            elif _contains(act,"not changed","continued","unchanged","increase"):
                df.at[i,"c_dechallenge"]="No change"
        # treatment change (Hartwig L1/L2) from action on dose
        if not str(row.get("s_treatment_change","")).strip() and has("action_on_dose"):
            act=str(row.get("action_on_dose","")).lower()
            if _contains(act,"withdraw","stopp","reduc","discontinu","held","changed"):
                df.at[i,"s_treatment_change"]="Drug held-stopped-changed"
            elif _contains(act,"not changed","continued","unchanged"):
                df.at[i,"s_treatment_change"]="No change"
        # severity flags from outcome / death_hosp / lifethreat columns
        out=str(row.get("adr_outcome","")).lower()
        dh = str(row.get("death_hosp","")).lower() if has("death_hosp") else ""
        lt = str(row.get("lifethreat_disab","")).lower() if has("lifethreat_disab") else ""
        if not str(row.get("s_contributed_death","")).strip():
            if _contains(out,"fatal","death","died") or _contains(dh,"death","died","fatal"):
                df.at[i,"s_contributed_death"]="Yes"
        if not str(row.get("s_permanent_harm","")).strip():
            if _contains(out,"sequelae") or _contains(lt,"disab","permanent"):
                df.at[i,"s_permanent_harm"]="Yes"
        if not str(row.get("s_intensive_care","")).strip():
            if _contains(lt,"life") or _contains(out,"life threat"):
                df.at[i,"s_intensive_care"]="Yes"
        if not str(row.get("s_caused_admission","")).strip():
            if _contains(dh,"hosp","admit","admission"):
                df.at[i,"s_caused_admission"]="Yes"
        if not str(row.get("s_increased_los","")).strip():
            if _contains(dh,"prolong","prolonged los","increase"):
                df.at[i,"s_increased_los"]="Yes"
        # if we had ANY severity source column, treat un-set flags as explicit "No"
        sev_source = any([out.strip(), dh.strip(), lt.strip(),
                          str(row.get("action_on_dose","")).strip()])
        if sev_source:
            for k in ["s_contributed_death","s_permanent_harm","s_intensive_care",
                      "s_caused_admission","s_increased_los","s_antidote_required"]:
                if not str(df.at[i,k]).strip() if k in df.columns else True:
                    if k not in df.columns: df[k]=""
                    if not str(df.at[i,k]).strip(): df.at[i,k]="No"
            if not str(df.at[i,"s_treatment_change"]).strip():
                df.at[i,"s_treatment_change"]="No change"
    return df

def g(row, key, default=""):
    v=row.get(key, default)
    if pd.isna(v): return default
    return str(v).strip()

# ---------------- NARANJO ----------------
def naranjo(row):
    miss=[]; s=0; known=0; detail=[]
    def kn(val,*ok): return 1 if val in ok else 0
    def add(label,val,pts):
        nonlocal s; s+=pts; detail.append(f"{label}={val or '?'} ({pts:+d})")
    v=g(row,"c_prev_reports"); add("prev-reports",v, 1 if v=="Yes" else 0); known+=kn(v,"Yes","No")
    v=g(row,"c_temporal")
    p = 2 if v in ("Plausible","Reasonable") else (-1 if v in ("Improbable","None") else 0)
    if v not in ("Plausible","Reasonable","Improbable","None"): miss.append("c_temporal")
    add("temporal",v,p); known+=kn(v,"Plausible","Reasonable","Improbable","None")
    v=g(row,"c_dechallenge"); p=1 if v=="Improved" else 0
    if v=="": miss.append("c_dechallenge")
    add("dechallenge",v,p); known+=kn(v,"Improved","No change","Worsened")
    v=g(row,"c_rechallenge"); p=2 if v=="Reaction recurred" else (-1 if v=="No recurrence" else 0)
    add("rechallenge",v,p); known+=kn(v,"Reaction recurred","No recurrence")
    v=g(row,"c_alt_causes")
    p = 2 if v=="Absent" else (-1 if v in ("Present","Possible") else 0)
    if v=="": miss.append("c_alt_causes")
    add("alt-causes",v,p); known+=kn(v,"Absent","Present","Possible")
    v=g(row,"c_placebo"); p=-1 if v=="Reaction present" else (1 if v=="Reaction absent" else 0)
    add("placebo",v,p); known+=kn(v,"Reaction present","Reaction absent")
    v=g(row,"c_toxic_level"); add("toxic-level",v,1 if v=="Yes" else 0); known+=kn(v,"Yes","No")
    v=g(row,"c_dose_response"); add("dose-response",v,1 if v=="Yes" else 0); known+=kn(v,"Yes","No")
    v=g(row,"c_prior_exposure"); add("prior-exposure",v,1 if v=="Yes" else 0); known+=kn(v,"Yes","No")
    v=g(row,"c_objective_confirm"); add("objective",v,1 if v=="Yes" else 0); known+=kn(v,"Yes","No")
    cat="Definite" if s>=9 else "Probable" if s>=5 else "Possible" if s>=1 else "Doubtful"
    return s, cat, miss, known, detail

# ---------------- WHO-UMC ----------------
def whoumc(row):
    # Faithful to the official WHO-UMC criteria (who.int). Returns (category, rationale).
    t=g(row,"c_temporal"); alt=g(row,"c_alt_causes"); de=g(row,"c_dechallenge"); defin=g(row,"c_event_definitive")
    if t=="":      return "Conditional / Unclassified","time relationship unknown - more data needed"
    if t=="None":  return "Unassessable / Unclassifiable","no temporal relationship"
    if t=="Improbable": return "Unlikely","time to onset makes relationship improbable"
    if t=="Plausible" and alt=="Absent" and de=="Improved" and defin=="Yes":
        return "Certain","plausible time + no alternative cause + positive dechallenge + definitive event"
    if alt=="Absent" and de=="Improved":
        return "Probable / Likely","reasonable time + alternative cause unlikely + reasonable dechallenge"
    return "Possible", f"reasonable time; alt-cause={alt or 'unknown'}, dechallenge={de or 'unknown/none'}"

# ---------------- HARTWIG & SIEGEL ----------------
def hartwig(row):
    keys=["s_contributed_death","s_permanent_harm","s_intensive_care","s_caused_admission",
          "s_increased_los","s_antidote_required","s_treatment_change"]
    miss=[] if [k for k in keys if g(row,k)!=""] else ["severity inputs"]
    def y(k): return g(row,k)=="Yes"
    if y("s_contributed_death"): return 7,"Severe",miss,"L7: reaction led to death"
    if y("s_permanent_harm"): return 6,"Severe",miss,"L6: permanent harm"
    if y("s_intensive_care"): return 5,"Severe",miss,"L5: required intensive care"
    if y("s_caused_admission") or y("s_increased_los"): return 4,"Moderate",miss,"L4: caused admission / prolonged stay"
    if y("s_antidote_required"): return 3,"Moderate",miss,"L3: antidote/treatment required"
    if g(row,"s_treatment_change")=="Drug held-stopped-changed": return 2,"Mild",miss,"L2: suspected drug held/changed"
    return 1,"Mild",miss,"L1: no change in treatment required"

# ---------------- SCHUMOCK & THORNTON ----------------
def preventability(row):
    A=["p_a1_allergy_history","p_a2_drug_inappropriate","p_a3_dose_inappropriate","p_a4_toxic_level","p_a5_known_treatment"]
    B=["p_b1_monitoring_missed","p_b2_interaction","p_b3_poor_compliance","p_b4_preventive_missed"]
    miss=[] if [k for k in A+B if g(row,k)!=""] else ["preventability inputs"]
    ay=[k.replace("p_a","A").split("_")[0] for k in A if g(row,k)=="Yes"]
    by=[k.replace("p_b","B").split("_")[0] for k in B if g(row,k)=="Yes"]
    if ay: return "Definitely preventable", miss, "Section A criteria met: "+", ".join(ay)
    if by: return "Probably preventable", miss, "Section B criteria met: "+", ".join(by)
    return "Not preventable", miss, "no preventability criterion met"

# ---------------- ICH E2A SERIOUSNESS ----------------
def ich_e2a(row):
    # ICH E2A: an event is SERIOUS if it meets ANY of 6 criteria. Seriousness (regulatory,
    # binary) is distinct from severity (Hartwig intensity). Returns (label, criteria-met).
    crit=[]
    out=g(row,"adr_outcome").lower(); dh=g(row,"death_hosp").lower(); lt=g(row,"lifethreat_disab").lower()
    desc=g(row,"adr_description").lower()
    if g(row,"s_contributed_death")=="Yes" or _contains(out,"fatal","death","died") or _contains(dh,"death","died","fatal"):
        crit.append("Death")
    if _contains(lt,"life") or g(row,"e2a_life_threatening").lower()=="yes" or _contains(out,"life threat"):
        crit.append("Life-threatening")
    if _contains(dh,"hosp","admit","admission","prolong") or g(row,"s_caused_admission")=="Yes" or g(row,"s_increased_los")=="Yes":
        crit.append("Hospitalisation / prolongation")
    if _contains(lt,"disab") or g(row,"s_permanent_harm")=="Yes" or _contains(out,"sequelae"):
        crit.append("Persistent / significant disability")
    if _contains(desc,"congenital","birth defect","anomaly") or g(row,"e2a_congenital_anomaly").lower()=="yes":
        crit.append("Congenital anomaly / birth defect")
    if g(row,"e2a_important_medical_event").lower()=="yes":
        crit.append("Important medical event")
    if crit:
        return "Serious", "; ".join(crit)
    if g(row,"seriousness").lower().startswith("yes"):
        return "Serious", "marked serious (specific criterion not recorded)"
    if g(row,"seriousness").lower().startswith("no") or out or dh or lt:
        return "Non-serious", "no ICH E2A seriousness criterion met"
    return "", "not assessable (no seriousness data)"

# ---------------- LOAD / SCORE / REPORT ----------------
def load_rows(path):
    if path.lower().endswith(".csv"):
        df=pd.read_csv(path, dtype=str, keep_default_na=False)
    else:
        xls=pd.ExcelFile(path)
        sheet=next((s for s in ["Upload template","Sheet1"] if s in xls.sheet_names), xls.sheet_names[0])
        df=pd.read_excel(path, sheet_name=sheet, dtype=str)
    df=normalize_columns(df)
    # keep genuine rows: must have a suspected drug
    if "suspected_drug" not in df.columns:
        raise SystemExit("No 'Drug'/'suspected_drug' column found after mapping.")
    drug=df["suspected_drug"].astype(str).str.strip()
    df=df[(drug!="") & (drug.str.lower()!="nan")].reset_index(drop=True)
    detect_date_format(df)
    df=derive_fields(df)
    return df

def score(df):
    out=[]
    ts=datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    for _,row in df.iterrows():
        ns,nc,nm,nknown,ndetail=naranjo(row)
        who,whoreason=whoumc(row)
        hl,hg,hm,hreason=hartwig(row)
        pv,pm,preason=preventability(row)
        ser,sercrit=ich_e2a(row)
        flags=[]
        if nm or who=="Conditional / Unclassified": flags.append("causality")
        if nknown<5: flags.append(f"causality low-data ({nknown}/10 Naranjo items known)")
        if hm: flags.append("severity")
        if pm: flags.append("preventability")
        integ=g(row,"data_integrity")
        if integ: flags.append("DATA INTEGRITY: "+integ)
        rec={
            "case_id":g(row,"case_id"),"patient_id":g(row,"patient_id"),
            "suspected_drug":g(row,"suspected_drug"),"adr_description":g(row,"adr_description"),
            "Naranjo score":ns,"Naranjo causality":nc,"Naranjo items known":f"{nknown}/10",
            "Naranjo breakdown":"; ".join(ndetail),
            "WHO-UMC causality":who,"WHO-UMC basis":whoreason,
            "Hartwig level":hl,"Severity grade":hg,"Severity basis":hreason,
            "Preventability":pv,"Preventability basis":preason,
            "Seriousness (ICH E2A)":ser,"Serious criteria":sercrit,
            "Data flags":"; ".join(flags) if flags else "complete",
            "Missing/assumed fields":"; ".join(sorted(set(nm+hm+pm))),
            "Assessed (UTC)":ts,
        }
        canon="|".join([rec["case_id"],rec["suspected_drug"],g(row,"adr_description"),
                        str(ns),nc,who,str(hl),hg,pv,ser,ts,CASPER_VERSION])
        rec["Row SHA-256"]=hashlib.sha256(canon.encode("utf-8")).hexdigest()
        out.append(rec)
    return pd.DataFrame(out)

def write_report(res, out_path):
    wb=Workbook(); thin=Side(style="thin",color="BFBFBF"); bord=Border(thin,thin,thin,thin)
    hf=Font(name=F,bold=True,color="FFFFFF"); hfill=PatternFill("solid",fgColor=BLUE)
    ws=wb.active; ws.title="Per-case results"; cols=list(res.columns)
    for j,c in enumerate(cols,1):
        cell=ws.cell(1,j,c); cell.font=hf; cell.fill=hfill; cell.border=bord
        cell.alignment=Alignment(wrap_text=True,horizontal="center",vertical="center")
    for i,(_,r) in enumerate(res.iterrows(),2):
        for j,c in enumerate(cols,1):
            cell=ws.cell(i,j,r[c]); cell.border=bord; cell.font=Font(name=F,size=10)
            cell.alignment=Alignment(vertical="center",wrap_text=True)
            if c=="Severity grade":
                cell.fill=PatternFill("solid",fgColor={"Mild":GREEN,"Moderate":AMBER,"Severe":RED}.get(r[c],"FFFFFF"))
            if c=="Seriousness (ICH E2A)" and r[c]=="Serious":
                cell.fill=PatternFill("solid",fgColor=RED)
            if c=="Data flags" and r[c]!="complete":
                cell.fill=PatternFill("solid",fgColor=AMBER)
    widthmap={"case_id":12,"patient_id":11,"suspected_drug":16,"adr_description":26,
        "Naranjo score":9,"Naranjo causality":14,"Naranjo items known":10,"Naranjo breakdown":46,
        "WHO-UMC causality":18,"WHO-UMC basis":40,"Hartwig level":8,"Severity grade":11,"Severity basis":30,
        "Preventability":18,"Preventability basis":30,"Seriousness (ICH E2A)":15,"Serious criteria":34,
        "Data flags":22,"Missing/assumed fields":28,"Assessed (UTC)":18,"Row SHA-256":30}
    for j,c in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(j)].width=widthmap.get(c,16)
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=40

    sm=wb.create_sheet("Summary")
    sm["A1"]="CASPER — Aggregate summary"; sm["A1"].font=Font(name=F,bold=True,size=14,color=NAVY)
    sm["A2"]=f"Total cases scored: {len(res)}"; sm["A2"].font=Font(name=F,bold=True)
    r=4
    def block(title, series, order=None):
        nonlocal r
        sm.cell(r,1,title).font=Font(name=F,bold=True,color="FFFFFF"); sm.cell(r,1).fill=PatternFill("solid",fgColor=NAVY)
        sm.cell(r,2).fill=PatternFill("solid",fgColor=NAVY); r+=1
        sm.cell(r,1,"Category").font=Font(name=F,bold=True); sm.cell(r,2,"n").font=Font(name=F,bold=True)
        sm.cell(r,3,"%").font=Font(name=F,bold=True); r+=1
        vc=series.value_counts(); cats=order if order else list(vc.index); start=r; tot=len(series)
        for cat in cats:
            n=int(vc.get(cat,0)); sm.cell(r,1,cat); sm.cell(r,2,n)
            sm.cell(r,3, f"{(100*n/tot):.1f}%" if tot else "0%"); r+=1
        ch=BarChart(); ch.type="bar"; ch.title=title; ch.height=5; ch.width=11
        ch.add_data(Reference(sm,min_col=2,min_row=start,max_row=r-1))
        ch.set_categories(Reference(sm,min_col=1,min_row=start,max_row=r-1)); ch.legend=None
        sm.add_chart(ch, f"E{start-1}"); r+=1
    block("Causality — Naranjo", res["Naranjo causality"], ["Definite","Probable","Possible","Doubtful"])
    block("Causality — WHO-UMC", res["WHO-UMC causality"])
    block("Severity — Hartwig & Siegel", res["Severity grade"], ["Mild","Moderate","Severe"])
    block("Preventability — Schumock & Thornton", res["Preventability"], ["Definitely preventable","Probably preventable","Not preventable"])
    sm.column_dimensions["A"].width=30; sm.column_dimensions["B"].width=8; sm.column_dimensions["C"].width=10

    fl=wb.create_sheet("Data flags"); flagged=res[res["Data flags"]!="complete"]
    fl["A1"]="Rows needing reviewer input (inputs missing or assumed)"; fl["A1"].font=Font(name=F,bold=True,size=12,color=NAVY)
    if len(flagged)==0:
        fl["A3"]="None — all rows had complete inputs."; fl["A3"].font=Font(name=F,italic=True)
    else:
        hdr=["case_id","suspected_drug","Data flags","Missing/assumed fields"]
        for j,c in enumerate(hdr,1):
            cell=fl.cell(3,j,c); cell.font=hf; cell.fill=hfill; cell.border=bord
        for i,(_,r2) in enumerate(flagged.iterrows(),4):
            for j,c in enumerate(hdr,1):
                cell=fl.cell(i,j,r2[c]); cell.border=bord; cell.font=Font(name=F,size=10); cell.alignment=Alignment(wrap_text=True,vertical="top")
        for j,w in enumerate([14,18,22,55],1): fl.column_dimensions[get_column_letter(j)].width=w
    cs=wb.create_sheet("Cite CASPER")
    cs.sheet_view.showGridLines=False
    cs["A1"]="If you use these results, please cite CASPER"
    cs["A1"].font=Font(name=F,bold=True,size=13,color=NAVY)
    cs["A3"]=CASPER_CITATION; cs["A3"].font=Font(name=F,size=11)
    cs["A5"]="Author: Puneet Paliwal   |   Licence: Apache-2.0   |   Version: "+CASPER_VERSION
    cs["A5"].font=Font(name=F,size=10,color="555555")
    cs["A7"]="Tool output is a research/decision-support aid; it does not replace clinical judgement."
    cs["A7"].font=Font(name=F,italic=True,size=10,color="555555")
    cs["A9"]="Audit & integrity"; cs["A9"].font=Font(name=F,bold=True,size=11,color=NAVY)
    cs["A10"]=("Each row carries an 'Assessed (UTC)' timestamp and a 'Row SHA-256' tamper-evidence hash "
               "computed over case id + drug + reaction + all scale outputs + timestamp + tool version. "
               "Re-hashing a row after any manual edit will not match the stored hash, revealing changes.")
    cs["A10"].font=Font(name=F,size=10,color="555555"); cs["A10"].alignment=Alignment(wrap_text=True)
    cs["A12"]="Scales: Naranjo + WHO-UMC (causality), Hartwig & Siegel (severity), Schumock & Thornton (preventability), ICH E2A (seriousness)."
    cs["A12"].font=Font(name=F,size=10,color="555555")
    cs.column_dimensions["A"].width=110
    # stamp a citation footer two rows under the per-case table
    last=ws.max_row+2
    ws.cell(last,1,"Generated by CASPER v"+CASPER_VERSION+" — please cite: "+CASPER_CITATION).font=Font(name=F,italic=True,size=9,color="777777")
    wb.save(out_path)

def main():
    if len(sys.argv)<2:
        print("Usage: python casper_engine.py <input.xlsx|csv> [output.xlsx]"); sys.exit(1)
    inp=sys.argv[1]
    out=sys.argv[2] if len(sys.argv)>2 else os.path.splitext(inp)[0].replace("_template","").replace("_upload","")+"_RESULTS.xlsx"
    df=load_rows(inp); res=score(df); write_report(res,out)
    print(f"Scored {len(res)} case(s).  ->  {out}")
    print(res[["case_id","Naranjo causality","WHO-UMC causality","Severity grade","Preventability","Seriousness (ICH E2A)","Data flags"]].to_string(index=False))

if __name__=="__main__":
    main()
